"""Export service for CSV, Excel, and PDF reports."""

import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.backtest import Backtest


class ExportService:
    @staticmethod
    def export_csv(backtest: Backtest) -> bytes:
        output = io.StringIO()
        writer = csv.writer(output)

        results = backtest.results or {}
        analytics = results.get("analytics", {})

        writer.writerow(["Equity Backtest Report"])
        writer.writerow(["Backtest ID", backtest.id])
        writer.writerow(["Name", backtest.name])
        writer.writerow(["Status", backtest.status])
        writer.writerow(["Created", str(backtest.created_at)])
        writer.writerow([])

        writer.writerow(["Performance Metrics"])
        for key, value in analytics.items():
            if key not in ("equity_curve", "drawdown_series"):
                writer.writerow([key, value])
        writer.writerow([])

        writer.writerow(["Equity Curve"])
        writer.writerow(["Date", "Value"])
        for point in analytics.get("equity_curve", []):
            writer.writerow([point["date"], point["value"]])
        writer.writerow([])

        writer.writerow(["Holdings"])
        writer.writerow(["Stock", "Weight", "Entry Date", "Exit Date", "Return %"])
        for h in results.get("holdings", []):
            writer.writerow([
                h.get("stock"), h.get("weight"), h.get("entry_date"),
                h.get("exit_date"), h.get("returns"),
            ])

        return output.getvalue().encode("utf-8")

    @staticmethod
    def export_excel(backtest: Backtest) -> bytes:
        wb = Workbook()
        results = backtest.results or {}
        analytics = results.get("analytics", {})

        ws_summary = wb.active
        ws_summary.title = "Summary"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        ws_summary["A1"] = "Equity Backtest Report"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary.append([])
        ws_summary.append(["Metric", "Value"])

        for key, value in analytics.items():
            if key not in ("equity_curve", "drawdown_series"):
                ws_summary.append([key.replace("_", " ").title(), value])

        ws_equity = wb.create_sheet("Equity Curve")
        ws_equity.append(["Date", "Portfolio Value"])
        ws_equity["A1"].font = header_font
        ws_equity["A1"].fill = header_fill
        ws_equity["B1"].font = header_font
        ws_equity["B1"].fill = header_fill
        for point in analytics.get("equity_curve", []):
            ws_equity.append([point["date"], point["value"]])

        ws_holdings = wb.create_sheet("Holdings")
        ws_holdings.append(["Stock", "Weight", "Entry Date", "Exit Date", "Return %"])
        for cell in ws_holdings[1]:
            cell.font = header_font
            cell.fill = header_fill
        for h in results.get("holdings", []):
            ws_holdings.append([
                h.get("stock"), h.get("weight"), h.get("entry_date"),
                h.get("exit_date"), h.get("returns"),
            ])

        ws_winners = wb.create_sheet("Top Winners")
        ws_winners.append(["Stock", "Return %", "Weight"])
        for w in results.get("top_winners", []):
            ws_winners.append([w.get("stock"), w.get("returns"), w.get("weight")])

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_pdf(backtest: Backtest) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5 * inch)
        styles = getSampleStyleSheet()
        story = []

        results = backtest.results or {}
        analytics = results.get("analytics", {})

        story.append(Paragraph("Equity Backtest Report", styles["Title"]))
        story.append(Paragraph(f"Backtest: {backtest.name or backtest.id}", styles["Heading2"]))
        story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
        story.append(Spacer(1, 12))

        metrics_data = [["Metric", "Value"]]
        for key in ["cagr", "total_return", "sharpe_ratio", "sortino_ratio", "max_drawdown", "volatility", "win_rate"]:
            if key in analytics:
                metrics_data.append([key.replace("_", " ").title(), f"{analytics[key]}"])

        metrics_table = Table(metrics_data, colWidths=[3 * inch, 2 * inch])
        metrics_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 20))

        if results.get("top_winners"):
            story.append(Paragraph("Top Winners", styles["Heading3"]))
            winners_data = [["Stock", "Return %"]]
            for w in results["top_winners"][:5]:
                winners_data.append([w.get("stock", ""), f"{w.get('returns', 0)}%"])
            winners_table = Table(winners_data, colWidths=[3 * inch, 2 * inch])
            winners_table.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#059669")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ]))
            story.append(winners_table)

        doc.build(story)
        return buffer.getvalue()
